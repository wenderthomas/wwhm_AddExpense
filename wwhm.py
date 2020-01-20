#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jan 14 22:03:07 2020

@author: WenderThomas
"""

import os
from datetime import date
<<<<<<< HEAD
from openpyxl import load_workbook, Workbook
=======
from openpyxl import load_workbook
>>>>>>> 13c5aa6f0989d5dfaa7a4ac3db8d1c850f257364

class AddExpense(object):
    """
    Add expenses to the selected .xlsx file and sort the rows by date
    
    args:
        file_name:  .xlsx file name
    
    return:
        None
    """
    
    def __init__(self, file_name):
        """Call all the methods in sequence"""
        self.file_name = file_name
        self.when = self.when()
        self.where = self.where()
        self.how_much = self.how_much()
        self.confirmation = self.confirm()
        self.save_func()
<<<<<<< HEAD
    
=======
>>>>>>> 13c5aa6f0989d5dfaa7a4ac3db8d1c850f257364
    
    #Funcao de entrada do valor da data da despesa
    def when(self):
        """Ask user when the expense was made"""
        while True:
            when_str = input ('When? ')
            try:
                when_list = when_str.split('/')
                if len(when_list[2]) != 4:                    
                    if len(when_list[2]) == 2:
                        when_list[2] = '20' + when_list[2]
                        self.when = date(int(when_list[2]), int(when_list[1]),\
                                         int(when_list[0]))
                        break
                    else:
                        continue
                self.when = date(int(when_list[2]), int(when_list[1]),\
                                 int(when_list[0]))
                break
            except:
                print ('Date format not recognized')
                print ('Type dd/mm/yyyy')
        return self.when
    
<<<<<<< HEAD
    
=======
>>>>>>> 13c5aa6f0989d5dfaa7a4ac3db8d1c850f257364
    #Funcao de entrada do valor da descricao da despesa
    def where(self):
        """Ask user where the expense was made"""
        self.where = input ('Where? ')
        return self.where
    
<<<<<<< HEAD
    
=======
>>>>>>> 13c5aa6f0989d5dfaa7a4ac3db8d1c850f257364
    #Funcao de entrada do valor da despesa
    def how_much(self):
        """Ask the user how much was spent"""
        while True:
            try:
                self.how_much = float(input ('How much? '))
                break
            except:
                print ('Amount not recognized')
        return self.how_much
    
<<<<<<< HEAD
    
    #Funcao de confirmacao dos dados digitados
    def confirm(self):
        """Ask the user to confirm if the data inserted is correct in order to 
        save"""
        while True:
            print ()
            print ('{:.2f} {:s} {:%d/%m/%Y}'.format(self.how_much,\
                                                   self.where,self.when))
            confirmation = input('Confirm? [Yes] or [No]: ')
            if confirmation[0].lower() == 'y' or \
            confirmation[0].lower() == 'n':
                return confirmation[0].lower()
            else:
                print('Wrong input')
                continue          
    
    
    #Funcao que salva os dados recebidos
    def save_func(self):
        """Save the data after if the user confirms"""
        sheet = self.when.strftime('%m') + '-' + str(self.when.year)
        self.file_exist = self.file_exist(sheet)
        if self.confirmation == 'y' and self.file_exist == 'created':
            workbook = load_workbook(filename=self.file_name)
            self.sheet_list = workbook.sheetnames
            if sheet in self.sheet_list:
                sheet_selected = workbook[sheet]                
            else:
                print('Sheet not found.')                
                print('Creating new sheet')
                workbook.create_sheet(sheet)
                sheet_selected = workbook[sheet]
                sheet_selected.append(['Descrição da despesa',\
                                       'Valor', 'Data'])
            sheet_selected.append([self.where, self.how_much, self.when])
            sheet_selected.auto_filter.ref = sheet_selected.dimensions
            print('Saving data to sheet {:s}'.format(sheet))                
            print()
            workbook.save(self.file_name)    
    
    
    #Funcao para verificar se o arquivo .xlsx existe na pasta
    def file_exist(self, sheet):
        """Check if the .xlsx file exists in the folder
        and if not creates it"""
        if not (os.path.exists(self.file_name) and\
                os.path.isfile(self.file_name)):
            print ('File not found')
            print ('Creating file {:s}'.format(self.file_name))
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = sheet
            worksheet = workbook[sheet]
            worksheet.append(['Descrição da despesa', 'Valor', 'Data'])
            workbook.save(self.file_name)
        return 'created'

    
# %% Run the code above in a loop
    
while True:    
    wwhm = AddExpense('Expenses.xlsx')
    print('Wish to add more expenses to the file?',end='')
    opt = input('Press [ENTER] to continue or N to stop')
    if opt.lower() == 'n':
        break
=======
    #Funcao de confirmacao dos dados digitados
    def confirm(self):
        """Ask the user to confirm if the data inserted is correct in order to 
        save"""
        while True:
            print ()
            print ('{:.2f} {:s} {:%d/%m/%Y}'.format(self.how_much,self.where,self.when))
            confirmation = input('Confirm? [Yes] or [No]: ')
            if confirmation[0].lower() == 'y' or \
            confirmation[0].lower() == 'n':
                return confirmation[0].lower()
            else:
                print('Wrong input')
                continue          
    
    #Funcao que salva os dados recebidos
    def save_func(self):
        """Save the data after if the user confirms"""
        #modificar para que o mes seja sempre com 2 digitos
        sheet = str(self.when.month) + '-' + str(self.when.year)
        #_______________________________________________________
        if self.confirmation == 'y':
            workbook = load_workbook(filename=self.file_name)
            self.sheet_list = workbook.sheetnames
            if sheet in self.sheet_list:
                sheet_selected = workbook[sheet]                
            else:
                print('Sheet not found.')                
                print('Creating new sheet')
                workbook.create_sheet(sheet)
                sheet_selected = workbook[sheet]
                sheet_selected.append(['Descrição da despesa',\
                                       'Valor', 'Data'])
            sheet_selected.append([self.where, self.how_much, self.when])
            sheet_selected.auto_filter.ref = sheet_selected.dimensions
            print('Saving data to sheet {:s}'.format(sheet))                
            workbook.save(self.file_name)    
    
# %% Run the code above
    
wwhm = AddExpense('Contabilidade.xlsx')
>>>>>>> 13c5aa6f0989d5dfaa7a4ac3db8d1c850f257364
